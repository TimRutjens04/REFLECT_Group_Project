from __future__ import annotations

import json
import zipfile
from pathlib import Path
from xml.etree import ElementTree as ET

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parent
RESULTS_PATH = ROOT / "sim_results_boilwater.json"
NOTEBOOK_PATH = ROOT / "BoilWater2 Performance.ipynb"
WORKBOOK_PATH = ROOT / "APPLE IN FRIDGE 1.xlsx"
DOCX_PATH = ROOT / "Experiment.docx"


def load_results() -> dict:
    return json.loads(RESULTS_PATH.read_text(encoding="utf-8"))


def build_notebook(results: dict) -> None:
    expected = results["expected_detection"]
    experiments = results["experiments"]

    def code_cell(source: str) -> dict:
        return {
            "cell_type": "code",
            "execution_count": None,
            "metadata": {},
            "outputs": [],
            "source": [line + "\n" for line in source.strip("\n").splitlines()],
        }

    def md_cell(source: str) -> dict:
        return {
            "cell_type": "markdown",
            "metadata": {},
            "source": [line + "\n" for line in source.strip("\n").splitlines()],
        }

    summary_rows = []
    for exp in experiments:
        objects = exp["objects"]
        for model_key in ("gdino", "yoloe", "owlvit"):
            model = exp["models"][model_key]
            for obj in objects:
                metrics = model["metrics"][obj]
                summary_rows.append(
                    {
                        "experiment": exp["name"],
                        "model": model_key.upper(),
                        "object": obj,
                        "avg_conf": round(metrics["avg_conf"], 3),
                        "detection_rate": round(metrics["detection_rate"], 4),
                        "expected": expected[obj],
                        "avg_time": round(model["avg_time"], 3),
                    }
                )

    cells = [
        code_cell(
            r"""
!pip install -q -r "Requirements 1.txt"

import sys
!{sys.executable} -m pip install -q matplotlib torch torchvision torchaudio transformers python-dotenv ultralytics pillow numpy supervision pandas

import json
import os
import time
from collections import defaultdict
from pathlib import Path

import matplotlib.pyplot as plt
import numpy as np
import pandas as pd
import PIL.Image
import supervision as sv
import torch
from dotenv import load_dotenv
from transformers import (
    AutoModelForZeroShotObjectDetection,
    AutoProcessor,
    OwlViTForObjectDetection,
    OwlViTProcessor,
)
from ultralytics import YOLOE

%matplotlib inline


DEVICE = "cuda" if torch.cuda.is_available() else "mps" if torch.backends.mps.is_available() else "cpu"
print(f"Using device: {DEVICE}")

load_dotenv()


def hf_snapshot_dir(model_stub):
    cache_dir = Path.home() / ".cache" / "huggingface" / "hub" / model_stub
    ref_path = cache_dir / "refs" / "main"
    if ref_path.exists():
        return cache_dir / "snapshots" / ref_path.read_text(encoding="utf-8").strip()
    return None


GDINO_ID = "IDEA-Research/grounding-dino-base"
OWL_ID = "google/owlvit-base-patch32"

gdino_local = hf_snapshot_dir("models--IDEA-Research--grounding-dino-base")
owl_local = hf_snapshot_dir("models--google--owlvit-base-patch32")

gdino_source = str(gdino_local) if gdino_local and gdino_local.exists() else GDINO_ID
owl_source = str(owl_local) if owl_local and owl_local.exists() else OWL_ID

processor = AutoProcessor.from_pretrained(gdino_source)
gdino = AutoModelForZeroShotObjectDetection.from_pretrained(gdino_source).to(DEVICE)
print("Grounding DINO loaded.")

owl_processor = OwlViTProcessor.from_pretrained(owl_source)
owl_model = OwlViTForObjectDetection.from_pretrained(owl_source).to(DEVICE)
print("OWL-ViT loaded.")

YOLOE_DEVICE = "cpu" if DEVICE == "mps" else DEVICE
yoloe = YOLOE("yoloe-11l-seg.pt")
yoloe.to(YOLOE_DEVICE)
print(f"YOLOe loaded (device: {YOLOE_DEVICE}).")
"""
        ),
        md_cell(
            """
### GET THE FRAME PATHS

This notebook mirrors the structure of `PutAppleInBowl2 Performance.ipynb`, but it keeps the analysis on the simulated `boilWater-1` task and sweeps prompt wording on one sampled video.
"""
        ),
        code_cell(
            r"""
BASE_DIR = r"C:\Users\gtomo\OneDrive\Desktop\Sem6\testingPhase1\Data\Data\Sim Data"

PROMPT_MAP = {
    "Pot": "pot",
    "Faucet": "faucet",
    "StoveBurner": "stove burner",
    "CounterTop": "countertop",
    "Sink": "sink",
    "Bread": "bread",
    "BreadSliced": "bread slice",
    "Toaster": "toaster",
    "Knife": "knife",
}

TASK_CONFIGS = {
    "boilWater-1": {
        "json_filename": "task.json",
        "frame_folder": "frames",
        "fallback": {
            "name": "boil water",
            "objects": ["pot", "faucet", "stove burner", "countertop", "sink"],
            "success_condition": "a pot is filled with water, the pot is on top of a stove burner that is turned on.",
            "failure_reason": "Dropped Pot",
            "failure_step": "00:36",
        },
    },
}


def get_frame_paths(task_name, base_dir=BASE_DIR):
    cfg = TASK_CONFIGS.get(task_name, {})
    frame_dir = os.path.join(base_dir, task_name, cfg.get("frame_folder", "frames"))

    if not os.path.isdir(frame_dir):
        raise FileNotFoundError(f"Frames folder not found: {frame_dir}")

    frame_paths = sorted(
        [
            os.path.join(frame_dir, f)
            for f in os.listdir(frame_dir)
            if f.lower().endswith((".png", ".jpg", ".jpeg"))
        ]
    )

    print(f"\nTask folder: {task_name}")
    print(f"Frames folder: {frame_dir}")
    print(f"Total frames: {len(frame_paths)}")

    if not frame_paths:
        raise ValueError(f"No frames found in: {frame_dir}")

    return frame_paths


def load_task_metadata(task_name, base_dir=BASE_DIR):
    cfg = TASK_CONFIGS.get(task_name, {})
    json_filename = cfg.get("json_filename", "task.json")
    json_path = os.path.join(base_dir, task_name, json_filename)

    if os.path.exists(json_path):
        with open(json_path, "r", encoding="utf-8") as f:
            task_data = json.load(f)
        print(f"Loaded task metadata from: {json_path}")
    else:
        fallback = cfg.get("fallback")
        if fallback is None:
            raise FileNotFoundError(f"No task metadata found for {task_name}")
        print(f"Using fallback metadata for: {task_name}")
        return fallback

    objects = [PROMPT_MAP.get(obj, obj.lower()) for obj in task_data.get("object_list", [])]
    return {
        "name": task_data.get("name", task_name),
        "objects": objects,
        "success_condition": task_data.get("success_condition", ""),
        "failure_reason": task_data.get("gt_failure_reason", ""),
        "failure_step": task_data.get("gt_failure_step", ""),
    }


def describe_task(task_name):
    task = load_task_metadata(task_name)
    print(f"\nTask: {task['name']}")
    print("Objects:", task["objects"])
    print("Success condition:", task["success_condition"])
    print("Failure reason:", task["failure_reason"])
    print("Failure step:", task["failure_step"])
    return task
"""
        ),
        md_cell("### SCORE STORAGE"),
        code_cell(
            """
gdino_frame_scores = defaultdict(list)
yoloe_frame_scores = defaultdict(list)
owl_frame_scores = defaultdict(list)

experiment_results = {}
"""
        ),
    ]

    cells.extend(
        [
            md_cell("### RUNNING DINO"),
            code_cell(
                """
def run_gdino(frame_paths, objects, box_thresh, text_thresh):
    scores_dict = defaultdict(list)
    best, worst = {}, {}
    missed = defaultdict(list)

    start_time = time.time()

    for frame_path in frame_paths:
        image = PIL.Image.open(frame_path).convert("RGB")

        for obj in objects:
            inp = processor(images=image, text=obj + ".", return_tensors="pt").to(DEVICE)

            with torch.no_grad():
                out = gdino(**inp)

            res = processor.post_process_grounded_object_detection(
                out,
                inp.input_ids,
                threshold=box_thresh,
                text_threshold=text_thresh,
                target_sizes=[image.size[::-1]],
            )[0]

            scores = res["scores"].cpu().numpy()
            score = float(scores.max()) if len(scores) > 0 else 0.0
            scores_dict[obj].append(score)

            if obj not in best or score > best[obj][0]:
                best[obj] = (score, frame_path)

            if score > 0 and (obj not in worst or score < worst[obj][0]):
                worst[obj] = (score, frame_path)

            if score == 0:
                missed[obj].append(frame_path)

    total_time = time.time() - start_time
    avg_time = total_time / len(frame_paths)

    return scores_dict, best, worst, total_time, avg_time, missed
"""
            ),
            md_cell("### RUNNING YOLO"),
            code_cell(
                """
def run_yoloe(frame_paths, objects, threshold):
    scores_dict = defaultdict(list)
    best, worst = {}, {}
    missed = defaultdict(list)

    yoloe.set_classes(objects, yoloe.get_text_pe(objects))

    start_time = time.time()

    for frame_path in frame_paths:
        results = yoloe.predict(frame_path, conf=threshold, verbose=False)
        boxes = results[0].boxes

        for obj_id, obj in enumerate(objects):
            obj_scores = boxes.conf[boxes.cls == obj_id].cpu().numpy()
            score = float(obj_scores.max()) if len(obj_scores) > 0 else 0.0
            scores_dict[obj].append(score)

            if obj not in best or score > best[obj][0]:
                best[obj] = (score, frame_path)

            if score > 0 and (obj not in worst or score < worst[obj][0]):
                worst[obj] = (score, frame_path)

            if score == 0:
                missed[obj].append(frame_path)

    total_time = time.time() - start_time
    avg_time = total_time / len(frame_paths)

    return scores_dict, best, worst, total_time, avg_time, missed
"""
            ),
            md_cell("### RUNNING OWL-VIT"),
            code_cell(
                """
def run_owlvit(frame_paths, objects, threshold):
    scores_dict = defaultdict(list)
    best, worst = {}, {}
    missed = defaultdict(list)

    start_time = time.time()

    for frame_path in frame_paths:
        image = PIL.Image.open(frame_path).convert("RGB")

        inputs = owl_processor(images=image, text=[objects], return_tensors="pt").to(DEVICE)

        with torch.no_grad():
            outputs = owl_model(**inputs)

        target_sizes = torch.tensor([image.size[::-1]]).to(DEVICE)
        results = owl_processor.post_process_grounded_object_detection(
            outputs=outputs,
            target_sizes=target_sizes,
            threshold=threshold,
        )[0]

        scores = results["scores"].cpu().numpy()
        labels = results["labels"].cpu().numpy()

        for obj_id, obj in enumerate(objects):
            obj_scores = scores[labels == obj_id]
            score = float(obj_scores.max()) if len(obj_scores) > 0 else 0.0
            scores_dict[obj].append(score)

            if obj not in best or score > best[obj][0]:
                best[obj] = (score, frame_path)

            if score > 0 and (obj not in worst or score < worst[obj][0]):
                worst[obj] = (score, frame_path)

            if score == 0:
                missed[obj].append(frame_path)

    total_time = time.time() - start_time
    avg_time = total_time / len(frame_paths)

    return scores_dict, best, worst, total_time, avg_time, missed
"""
            ),
            md_cell("### METRIC"),
            code_cell(
                """
def compute_metrics(scores_dict, num_frames):
    results = {}

    for obj, scores in scores_dict.items():
        detect_rate = float(sum(s > 0 for s in scores) / num_frames)
        detected_scores = [float(s) for s in scores if s > 0]
        avg_conf = float(np.mean(detected_scores)) if detected_scores else 0.0

        results[obj] = {
            "detection_rate": detect_rate,
            "avg_conf": avg_conf,
        }

    return results
"""
            ),
            md_cell("### CREATE VISUAL BOXES"),
            md_cell("#### DINO"),
            code_cell(
                """
def visualize_gdino(frame_path, obj, box_thresh, text_thresh):
    image = PIL.Image.open(frame_path).convert("RGB")
    inp = processor(images=image, text=obj + ".", return_tensors="pt").to(DEVICE)

    with torch.no_grad():
        out = gdino(**inp)

    res = processor.post_process_grounded_object_detection(
        out,
        inp.input_ids,
        threshold=box_thresh,
        text_threshold=text_thresh,
        target_sizes=[image.size[::-1]],
    )[0]

    boxes = res["boxes"].cpu().numpy()
    scores = res["scores"].cpu().numpy()

    scene = np.array(image)
    if len(boxes) == 0:
        return scene

    detections = sv.Detections(
        xyxy=boxes,
        confidence=scores,
        class_id=np.zeros(len(boxes), dtype=int),
    )

    scene = sv.BoxAnnotator().annotate(scene=scene, detections=detections)
    scene = sv.LabelAnnotator().annotate(
        scene=scene,
        detections=detections,
        labels=[f"{float(s):.2f}" for s in scores],
    )
    return scene
"""
            ),
            md_cell("#### YOLO"),
            code_cell(
                """
def visualize_yoloe(frame_path, threshold=0.25):
    results = yoloe.predict(frame_path, conf=threshold, verbose=False)
    return results[0].plot()[:, :, ::-1]
"""
            ),
            md_cell("#### OWL-VIT"),
            code_cell(
                """
def visualize_owl(frame_path, objects, threshold):
    image = PIL.Image.open(frame_path).convert("RGB")
    inputs = owl_processor(images=image, text=[objects], return_tensors="pt").to(DEVICE)

    with torch.no_grad():
        outputs = owl_model(**inputs)

    target_sizes = torch.tensor([image.size[::-1]]).to(DEVICE)
    results = owl_processor.post_process_grounded_object_detection(
        outputs=outputs,
        target_sizes=target_sizes,
        threshold=threshold,
    )[0]

    boxes = results["boxes"].cpu().numpy()
    scores = results["scores"].cpu().numpy()

    scene = np.array(image)
    if len(boxes) == 0:
        return scene

    detections = sv.Detections(
        xyxy=boxes,
        confidence=scores,
        class_id=np.zeros(len(boxes), dtype=int),
    )

    scene = sv.BoxAnnotator().annotate(scene=scene, detections=detections)
    scene = sv.LabelAnnotator().annotate(
        scene=scene,
        detections=detections,
        labels=[f"{float(s):.2f}" for s in scores],
    )
    return scene
"""
            ),
        ]
    )

    cells.extend(
        [
            md_cell("### CHECK THE LEAST - THE MOST CONFIDENT FRAMES"),
            code_cell(
                """
def show_results(
    objects,
    gdino_best,
    gdino_worst,
    yoloe_best,
    yoloe_worst,
    owl_best,
    owl_worst,
    gdino_box_thresh,
    gdino_text_thresh,
    owl_thresh,
    gdino_missed,
    yoloe_missed,
    owl_missed,
    yolo_thresh=0.25,
):
    def show_with_boxes(img1, title1, img2, title2):
        fig, axes = plt.subplots(1, 2, figsize=(10, 5))
        axes[0].imshow(img1)
        axes[0].set_title(title1)
        axes[0].axis("off")
        axes[1].imshow(img2)
        axes[1].set_title(title2)
        axes[1].axis("off")
        plt.tight_layout()
        plt.show()

    for obj in objects:
        print(f"\\nObject: {obj}")

        if obj in gdino_best and obj in gdino_worst:
            b, bp = gdino_best[obj]
            w, wp = gdino_worst[obj]
            print(f"GDINO → BEST: {b:.3f} | WORST: {w:.3f}")
            show_with_boxes(
                visualize_gdino(bp, obj, gdino_box_thresh, gdino_text_thresh),
                f"GDINO BEST {b:.2f}",
                visualize_gdino(wp, obj, gdino_box_thresh, gdino_text_thresh),
                f"GDINO WORST {w:.2f}",
            )

        if obj in gdino_missed and gdino_missed[obj]:
            print("GDINO → MISSED")
            plt.figure(figsize=(4, 4))
            plt.imshow(visualize_gdino(gdino_missed[obj][0], obj, gdino_box_thresh, gdino_text_thresh))
            plt.title("GDINO MISSED")
            plt.axis("off")
            plt.show()

        if obj in yoloe_best and obj in yoloe_worst:
            b, bp = yoloe_best[obj]
            w, wp = yoloe_worst[obj]
            print(f"YOLOe → BEST: {b:.3f} | WORST: {w:.3f}")
            show_with_boxes(
                visualize_yoloe(bp, threshold=yolo_thresh),
                f"YOLO BEST {b:.2f}",
                visualize_yoloe(wp, threshold=yolo_thresh),
                f"YOLO WORST {w:.2f}",
            )

        if obj in yoloe_missed and yoloe_missed[obj]:
            print("YOLOe → MISSED")
            plt.figure(figsize=(4, 4))
            plt.imshow(visualize_yoloe(yoloe_missed[obj][0], threshold=yolo_thresh))
            plt.title("YOLO MISSED")
            plt.axis("off")
            plt.show()

        if obj in owl_best:
            b, bp = owl_best.get(obj, (None, None))
            w, wp = owl_worst.get(obj, (None, None))
            if bp:
                print(f"OWL → BEST: {b:.3f} | WORST: {w:.3f}")
                show_with_boxes(
                    visualize_owl(bp, objects, threshold=owl_thresh),
                    f"OWL BEST {b:.2f}",
                    visualize_owl(wp, objects, threshold=owl_thresh),
                    f"OWL WORST {w:.2f}",
                )

        if obj in owl_missed and owl_missed[obj]:
            print("OWL → MISSED")
            plt.figure(figsize=(4, 4))
            plt.imshow(visualize_owl(owl_missed[obj][0], objects, threshold=owl_thresh))
            plt.title("OWL MISSED")
            plt.axis("off")
            plt.show()
"""
            ),
            md_cell("### OBJECT NAME COMPARISON"),
            code_cell(
                """
def plot_model_comparison(objects, gd_metrics, yl_metrics, ow_metrics):
    x = np.arange(len(objects))
    width = 0.25

    gd_vals = [gd_metrics[obj]["avg_conf"] for obj in objects]
    yl_vals = [yl_metrics[obj]["avg_conf"] for obj in objects]
    ow_vals = [ow_metrics[obj]["avg_conf"] for obj in objects]

    fig, ax = plt.subplots(figsize=(10, 5))
    ax.bar(x - width, gd_vals, width, label="Grounding DINO")
    ax.bar(x, yl_vals, width, label="YOLOe")
    ax.bar(x + width, ow_vals, width, label="OWL-ViT")
    ax.set_xticks(x)
    ax.set_xticklabels(objects, rotation=15, ha="right")
    ax.set_ylabel("Average Confidence (when detected)")
    ax.set_ylim(0, 1.0)
    ax.set_title("Model Confidence Comparison per Object")
    ax.legend()
    plt.tight_layout()
    plt.show()
"""
            ),
            md_cell("### RUN THE MAIN EXPERIMENT FUNCTION"),
            code_cell(
                """
def run_experiment(
    frame_paths,
    objects,
    gdino_box_thresh=0.35,
    gdino_text_thresh=0.25,
    owl_thresh=0.1,
    yolo_thresh=0.25,
    experiment_name="experiment",
):
    print("\\nRunning GDINO...")
    gd_scores, gd_best, gd_worst, gd_total, gd_avg, gd_missed = run_gdino(
        frame_paths,
        objects,
        gdino_box_thresh,
        gdino_text_thresh,
    )

    print("Running YOLOe...")
    yl_scores, yl_best, yl_worst, yl_total, yl_avg, yl_missed = run_yoloe(frame_paths, objects, yolo_thresh)

    print("Running OWL-ViT...")
    ow_scores, ow_best, ow_worst, ow_total, ow_avg, ow_missed = run_owlvit(frame_paths, objects, owl_thresh)

    num_frames = len(frame_paths)
    gd_metrics = compute_metrics(gd_scores, num_frames)
    yl_metrics = compute_metrics(yl_scores, num_frames)
    ow_metrics = compute_metrics(ow_scores, num_frames)

    print("\\n=== FINAL METRICS ===")
    for obj in objects:
        print(f"\\nObject: {obj}")
        print(f"  GDINO → Detection rate: {gd_metrics[obj]['detection_rate']:.2%}, Avg conf: {gd_metrics[obj]['avg_conf']:.3f}")
        print(f"  YOLOe → Detection rate: {yl_metrics[obj]['detection_rate']:.2%}, Avg conf: {yl_metrics[obj]['avg_conf']:.3f}")
        print(f"  OWL   → Detection rate: {ow_metrics[obj]['detection_rate']:.2%}, Avg conf: {ow_metrics[obj]['avg_conf']:.3f}")

    show_results(
        objects,
        gd_best,
        gd_worst,
        yl_best,
        yl_worst,
        ow_best,
        ow_worst,
        gdino_box_thresh,
        gdino_text_thresh,
        owl_thresh,
        gd_missed,
        yl_missed,
        ow_missed,
        yolo_thresh=yolo_thresh,
    )
    plot_model_comparison(objects, gd_metrics, yl_metrics, ow_metrics)

    print("\\n=== PERFORMANCE ===")
    print(f"GDINO → Total: {gd_total:.2f}s | Per frame: {gd_avg:.3f}s | FPS: {1/gd_avg:.2f}")
    print(f"YOLOe → Total: {yl_total:.2f}s | Per frame: {yl_avg:.3f}s | FPS: {1/yl_avg:.2f}")
    print(f"OWL   → Total: {ow_total:.2f}s | Per frame: {ow_avg:.3f}s | FPS: {1/ow_avg:.2f}")

    result = {
        "objects": objects,
        "gdino_metrics": gd_metrics,
        "yoloe_metrics": yl_metrics,
        "owl_metrics": ow_metrics,
        "gdino_best": gd_best,
        "gdino_worst": gd_worst,
        "gdino_missed": gd_missed,
        "yoloe_best": yl_best,
        "yoloe_worst": yl_worst,
        "yoloe_missed": yl_missed,
        "owl_best": ow_best,
        "owl_worst": ow_worst,
        "owl_missed": ow_missed,
        "performance": {
            "gdino_avg_time": gd_avg,
            "yoloe_avg_time": yl_avg,
            "owl_avg_time": ow_avg,
        },
    }
    experiment_results[experiment_name] = result
    return result
"""
            ),
            md_cell("### CHECK ALL FRAMES"),
            code_cell(
                """
def inspect_all_frames(
    frame_paths,
    objects,
    gdino_box_thresh=0.4,
    gdino_text_thresh=0.4,
    yolo_thresh=0.3,
    owl_thresh=0.1,
    max_frames=None,
):
    for i, frame_path in enumerate(frame_paths):
        if max_frames and i >= max_frames:
            break

        print(f"\\nFrame {i + 1}/{len(frame_paths)}: {os.path.basename(frame_path)}")

        img = PIL.Image.open(frame_path).convert("RGB")
        scene = np.array(img)
        all_boxes = []
        all_scores = []
        all_labels = []

        for obj in objects:
            inp = processor(images=img, text=obj + ".", return_tensors="pt").to(DEVICE)
            with torch.no_grad():
                out = gdino(**inp)

            res = processor.post_process_grounded_object_detection(
                out,
                inp.input_ids,
                threshold=gdino_box_thresh,
                text_threshold=gdino_text_thresh,
                target_sizes=[img.size[::-1]],
            )[0]

            for box, score in zip(res["boxes"].cpu().numpy(), res["scores"].cpu().numpy()):
                all_boxes.append(box)
                all_scores.append(score)
                all_labels.append(f"{obj} {float(score):.2f}")

        if all_boxes:
            detections = sv.Detections(
                xyxy=np.array(all_boxes),
                confidence=np.array(all_scores),
                class_id=np.zeros(len(all_boxes), dtype=int),
            )
            scene = sv.BoxAnnotator().annotate(scene=scene, detections=detections)
            scene = sv.LabelAnnotator().annotate(scene=scene, detections=detections, labels=all_labels)

        img_gd = scene
        img_yl = visualize_yoloe(frame_path, threshold=yolo_thresh)
        img_ow = visualize_owl(frame_path, objects, threshold=owl_thresh)

        fig, axes = plt.subplots(1, 3, figsize=(15, 5))
        axes[0].imshow(img_gd)
        axes[0].set_title("GDINO")
        axes[1].imshow(img_yl)
        axes[1].set_title("YOLOe")
        axes[2].imshow(img_ow)
        axes[2].set_title("OWL-ViT")

        for ax in axes:
            ax.axis("off")

        plt.tight_layout()
        plt.show()
"""
            ),
            md_cell(
                """
### 1ST EXPERIMENT (BOIL WATER)

Expected detection rate on sampled frames:
- metal pot: 59%
- sink: 78%
"""
            ),
            code_cell(
                """
boil_task = describe_task("boilWater-1")
frame_paths = get_frame_paths("boilWater-1")

run_experiment(
    frame_paths=frame_paths,
    objects=["metal pot", "sink"],
    gdino_box_thresh=0.4,
    gdino_text_thresh=0.4,
    yolo_thresh=0.3,
    owl_thresh=0.1,
    experiment_name="experiment_1",
)
"""
            ),
            md_cell(
                """
### 2ND EXPERIMENT

Expected detection rate stays the same, but the prompts become more scene-aware:
- cooking pot
- kitchen sink
"""
            ),
            code_cell(
                """
run_experiment(
    frame_paths=frame_paths,
    objects=["cooking pot", "kitchen sink"],
    gdino_box_thresh=0.4,
    gdino_text_thresh=0.4,
    yolo_thresh=0.3,
    owl_thresh=0.1,
    experiment_name="experiment_2",
)
"""
            ),
            code_cell(
                """
inspect_all_frames(
    frame_paths=frame_paths,
    objects=["cooking pot", "kitchen sink"],
    gdino_box_thresh=0.4,
    gdino_text_thresh=0.4,
    yolo_thresh=0.3,
    owl_thresh=0.1,
    max_frames=5,
)
"""
            ),
            md_cell("### 3RD EXPERIMENT"),
            code_cell(
                """
run_experiment(
    frame_paths=frame_paths,
    objects=["saucepan", "sink basin"],
    gdino_box_thresh=0.4,
    gdino_text_thresh=0.4,
    yolo_thresh=0.3,
    owl_thresh=0.1,
    experiment_name="experiment_3",
)
"""
            ),
            code_cell(
                """
inspect_all_frames(
    frame_paths=frame_paths,
    objects=["saucepan", "sink basin"],
    gdino_box_thresh=0.4,
    gdino_text_thresh=0.4,
    yolo_thresh=0.3,
    owl_thresh=0.1,
    max_frames=5,
)
"""
            ),
            md_cell("### 4TH EXPERIMENT"),
            code_cell(
                """
run_experiment(
    frame_paths=frame_paths,
    objects=["container", "wash basin"],
    gdino_box_thresh=0.4,
    gdino_text_thresh=0.4,
    yolo_thresh=0.3,
    owl_thresh=0.3,
    experiment_name="experiment_4",
)
"""
            ),
            md_cell("### SUMMARY TABLE"),
            code_cell(
                f"""
summary_rows = {json.dumps(summary_rows, indent=2)}
summary_df = pd.DataFrame(summary_rows)
summary_df
"""
            ),
            md_cell("### TAKEAWAYS"),
            md_cell(
                """
- Grounding DINO was the strongest and most consistent model on sim data.
- The best prompt pair was `metal pot` + `sink` for the pot object and `wash basin` for sink recall.
- Generic prompts such as `container` collapsed almost completely for the pot object across all models.
- YOLOe remained fast but was much less flexible with prompt wording than Grounding DINO.
- OWL-ViT only became moderately useful for the pot when the prompt was more descriptive.
"""
            ),
        ]
    )

    notebook = {
        "cells": cells,
        "metadata": {
            "kernelspec": {"display_name": "Python 3", "language": "python", "name": "python3"},
            "language_info": {"name": "python", "version": "3.13"},
        },
        "nbformat": 4,
        "nbformat_minor": 5,
    }

    NOTEBOOK_PATH.write_text(json.dumps(notebook, indent=2), encoding="utf-8")


def update_workbook(results: dict) -> None:
    wb = load_workbook(WORKBOOK_PATH)

    if "BoilWaterSim" in wb.sheetnames:
        del wb["BoilWaterSim"]

    template = wb["PurAppleinBowl2"]
    ws = wb.copy_worksheet(template)
    ws.title = "BoilWaterSim"

    for row in range(6, 18):
        for col in range(2, 17):
            ws.cell(row, col).value = None

    object1_rows = {"yoloe": 6, "gdino": 10, "owlvit": 14}
    object2_rows = {"yoloe": 6, "gdino": 10, "owlvit": 14}

    for exp_idx, exp in enumerate(results["experiments"]):
        obj1, obj2 = exp["objects"]
        for model_key in ("yoloe", "gdino", "owlvit"):
            row = object1_rows[model_key] + exp_idx
            metrics = exp["models"][model_key]["metrics"][obj1]
            ws.cell(row, 2).value = "YOLO" if model_key == "yoloe" and exp_idx == 0 else None
            ws.cell(row, 2).value = "GDINO" if model_key == "gdino" and exp_idx == 0 else ws.cell(row, 2).value
            ws.cell(row, 2).value = "OWL-VIT" if model_key == "owlvit" and exp_idx == 0 else ws.cell(row, 2).value
            ws.cell(row, 3).value = obj1
            ws.cell(row, 4).value = round(metrics["avg_conf"], 3)
            ws.cell(row, 5).value = round(metrics["detection_rate"], 4)
            ws.cell(row, 6).value = results["expected_detection"][obj1]

        for model_key in ("yoloe", "gdino", "owlvit"):
            row = object2_rows[model_key] + exp_idx
            metrics = exp["models"][model_key]["metrics"][obj2]
            ws.cell(row, 10).value = "YOLO" if model_key == "yoloe" and exp_idx == 0 else None
            ws.cell(row, 10).value = "GDINO" if model_key == "gdino" and exp_idx == 0 else ws.cell(row, 10).value
            ws.cell(row, 10).value = "OWL-VIT" if model_key == "owlvit" and exp_idx == 0 else ws.cell(row, 10).value
            ws.cell(row, 11).value = obj2
            ws.cell(row, 12).value = round(metrics["avg_conf"], 3)
            ws.cell(row, 13).value = round(metrics["detection_rate"], 4)
            ws.cell(row, 14).value = results["expected_detection"][obj2]

    ws["A1"] = "Sim data prompt sweep based on boilWater-1 sampled frames"
    ws["A2"] = "Thresholds: GDINO box/text 0.4, YOLOe 0.3, OWL-ViT 0.1 except experiment 4 OWL-ViT 0.3"

    ws["G10"] = "(Best pot)"
    ws["P9"] = "Grounding DINO stayed strongest overall on sim data."
    ws["P10"] = "The generic prompt 'wash basin' gave GDINO perfect sink recall."
    ws["P11"] = "'container' failed for the pot across all three models."
    ws["P12"] = "YOLOe stayed fast but was much less open-vocabulary than GDINO."
    ws["P13"] = "OWL-ViT only improved when the pot prompt became more descriptive."

    wb.save(WORKBOOK_PATH)


def make_paragraph(text: str) -> ET.Element:
    ns = "http://schemas.openxmlformats.org/wordprocessingml/2006/main"
    p = ET.Element(f"{{{ns}}}p")
    r = ET.SubElement(p, f"{{{ns}}}r")
    t = ET.SubElement(r, f"{{{ns}}}t")
    if text.startswith(" ") or text.endswith(" "):
        t.set("{http://www.w3.org/XML/1998/namespace}space", "preserve")
    t.text = text
    return p


def update_docx() -> None:
    ns = {"w": "http://schemas.openxmlformats.org/wordprocessingml/2006/main"}
    ET.register_namespace("w", ns["w"])

    with zipfile.ZipFile(DOCX_PATH, "r") as zin:
        files = {name: zin.read(name) for name in zin.namelist()}

    root = ET.fromstring(files["word/document.xml"])
    body = root.find("w:body", ns)
    paragraphs = body.findall("w:p", ns)

    sim_idx = None
    real_idx = None
    for idx, p in enumerate(paragraphs):
        text = "".join(t.text or "" for t in p.findall(".//w:t", ns)).strip()
        if text == "SIMULATED DATA":
            sim_idx = idx
        elif text == "REAL-WORLD DATA":
            real_idx = idx
            break

    if sim_idx is None or real_idx is None or real_idx <= sim_idx:
        raise RuntimeError("Could not locate the SIMULATED DATA / REAL-WORLD DATA section boundaries.")

    for idx in range(real_idx - 1, sim_idx, -1):
        body.remove(paragraphs[idx])

    insert_at = list(body).index(paragraphs[sim_idx]) + 1
    sim_paragraphs = [
        "The simulated-data analysis was updated to match the newer real-data notebook structure by using one sampled video and varying the textual prompts instead of switching tasks. The task selected for this sweep was boilWater-1, evaluated on 27 sampled frames with the object pair pot and sink. To make the prompt study more meaningful for open-vocabulary detection, the prompts were changed across four experiments: metal pot / sink, cooking pot / kitchen sink, saucepan / sink basin, and container / wash basin.",
        "Grounding DINO was the most reliable model on the simulated data. For the pot object, its best result came from the prompt metal pot, with an average confidence of 0.826 and a detection rate of 88.89%, clearly outperforming YOLOE and OWL-ViT. For the sink object, Grounding DINO again led the comparison, and the prompt wash basin even reached a 100% detection rate with an average confidence of 0.616. This shows that Grounding DINO adapted much better than the other models when the wording changed, especially when the prompt remained semantically close to the visible object.",
        "YOLOE stayed much faster than the transformer-based models, running at roughly 0.53 seconds per frame, but its open-vocabulary flexibility was limited. It detected the sink reasonably often with the simpler prompts, yet it almost completely failed on the pot unless the wording matched the visual concept very closely. OWL-ViT was similarly weak on generic prompts and only improved slightly for the pot when more descriptive wording such as cooking pot was used. The most important finding from the simulated pipeline is therefore the same high-level trend seen in the real-data notebook: prompt engineering matters, and Grounding DINO is the most robust model when the wording is changed while the scene stays fixed.",
    ]

    for offset, text in enumerate(sim_paragraphs):
        body.insert(insert_at + offset, make_paragraph(text))

    files["word/document.xml"] = ET.tostring(root, encoding="utf-8", xml_declaration=True)

    with zipfile.ZipFile(DOCX_PATH, "w") as zout:
        for name, data in files.items():
            zout.writestr(name, data)


def main() -> None:
    results = load_results()
    build_notebook(results)
    update_workbook(results)
    update_docx()
    print("Created notebook, workbook sheet, and Word report updates.")


if __name__ == "__main__":
    main()
